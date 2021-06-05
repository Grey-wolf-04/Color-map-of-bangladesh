import pandas as pd
from PIL import ImageTk,Image
import math
import os
import time
import matplotlib.pyplot as plt

from tkinter import *
import os

from matplotlib.widgets import Button
from openpyxl.workbook import *
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import colorchooser
from tkinter import ttk

import cv2
import numpy as np
from PIL import ImageTk,Image


root=Tk()
my_notebook=ttk.Notebook(root)
my_notebook.pack(pady=0)

frame1=Frame(my_notebook,width=50,height=50)
frame2=Frame(my_notebook,width=50,height=50)


frame1.pack(fill='both',expand=1)
frame2.pack(fill='both',expand=1)

my_notebook.add(frame1,text='new map')
my_notebook.add(frame2,text='load map')

import matplotlib.image as mpimg

global img



###############################################################################################



rgb_mega=[[]]

increment=25

for i in range(0,6):
	if i==0:
		r=250
		g=40
		b=70
		tr=[]
		for j in range(0,8):
			
			tr.append([r,g+j*increment,b+j*increment])
			
		rgb_mega.append(tr)

	if i==1:
		r=37
		g=33
		b=164
		tr=[]
		for j in range(0,8):
			
			tr.append([r+j*increment,g+j*increment,b])
			
		rgb_mega.append(tr)
	if i==2:
		r=40
		g=233
		b=35
		tr=[]
		for j in range(0,8):
			
			tr.append([r+j*increment,g,b+j*increment])
			
		rgb_mega.append(tr)
	if i==3:
		r=240
		g=240
		b=35
		tr=[]
		for j in range(0,8):
			
			tr.append([r,g,b+j*increment])
			
		rgb_mega.append(tr)
	if i==4:
		r=30
		g=192
		b=203
		tr=[]
		for j in range(0,8):
			
			tr.append([r+j*increment,g,b])
			
		rgb_mega.append(tr)

	if i==5:
		r=255
		g=50
		b=16
		tr=[]
		for j in range(0,8):
			
			tr.append([r,g+j*increment,b+j*increment])
			
		rgb_mega.append(tr)

my_canvas=Canvas(frame1)
my_canvas.pack(fill='both',expand=1)
text_label=Label(frame1,text='Use arrow keys to change')
text_label.place(x=20,y=25)





class bd_division:
	def __init__(self,name,position,val,rank,rgb):
		self.name=name
		self.position=position
		self.val=val
		self.rank=rank
		self.rgb=rgb
div_objects=[]
rp=bd_division('Rangpur',[126,107],0,0,[129,227,86])
div_objects.append(rp)
rj=bd_division('Rajshahi',[120,244],0,0,[134,220,166])
div_objects.append(rj)
my=bd_division('Mymensingh',[250,202],0,0,[91,203,255])
div_objects.append(my)
sy=bd_division('Sylhet',[381,241],0,0,[123,234,5])
div_objects.append(sy)
dk=bd_division('Dhaka',[248,213],0,0,[178,142,199])
div_objects.append(dk)
kh=bd_division('Khulna',[151,416],0,0,[140,213,247])
div_objects.append(kh)
br=bd_division('Barisal',[264,466],0,0,[255,0,102])
div_objects.append(br)
ch=bd_division('Chittagong',[377,459],0,0,[217,71,150])
div_objects.append(ch)

def division_2(index,color):
	def convert():
		pass


	print(index,color)

	for i in range(0,len(div_objects)):
	
		df=pd.read_csv(f"{div_objects[i].name}/trend/{div_objects[i].name}_trend.csv")
		
		try:
			os.remove(f"{div_objects[i].name}/trend/{div_objects[i].name}_trend.xlsx")
		except:
			pass
		writer=pd.ExcelWriter(f"{div_objects[i].name}/trend/{div_objects[i].name}_trend.xlsx")
		df.to_excel(writer,index=False)
		writer.save()
		wb=load_workbook(f"{div_objects[i].name}/trend/{div_objects[i].name}_trend.xlsx")
		sheet=wb['Sheet1']
		for j in range(2,len(sheet['A'])+1):
			#print(sheet.cell(row=j,column=6).value)
			if sheet.cell(row=j,column=3).value==index.lower():
				div_objects[i].val=sheet.cell(row=j,column=6).value
	trr=[]
	for i in range(0,len(div_objects)):
		trr.append(div_objects[i].val)
	trr.sort()
	print(trr)
	for i in range(0,8):
		for j in range(0,8):
			if trr[i]==div_objects[j].val:
				div_objects[j].rank=i+1
	
	img=cv2.imread(f'bd.png')
	img=cv2.cvtColor(img,cv2.COLOR_BGR2RGB)

	for i in range(0,8):
		img[np.where((img==div_objects[i].rgb).all(axis=2))]=rgb_mega[color][div_objects[i].rank-1]
	cv2.putText(img,f'{index}',(220,50),cv2.FONT_HERSHEY_TRIPLEX,0.5,(255,0,0),1)	
	cv2.putText(img,f'{rp.val}',(150,525),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)
	cv2.putText(img,f'{rj.val}',(150,535),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)
	cv2.putText(img,f'{dk.val}',(150,545),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)
	cv2.putText(img,f'{kh.val}',(150,557),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)
	cv2.putText(img,f'{br.val}',(150,570),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)
	cv2.putText(img,f'{ch.val}',(150,580),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)
	cv2.putText(img,f'{sy.val}',(150,595),cv2.FONT_HERSHEY_TRIPLEX,0.35,(255,0,0),1)	


	img=cv2.cvtColor(img,cv2.COLOR_RGB2BGR)
	cv2.imwrite(f'bd_2.png',img)
	cv2.imwrite(f'py_plots/{color}_{index}_All_plots_in_color.png',img)
	try:
		bd_2=PhotoImage(file='bd_2.png')
		i = Image.open('bd_2.png')
		
		iar = np.asarray(i)
		fig=plt.imshow(iar)
		plt.xlabel(f'{index} For Year 1988-2018')
		plt.show()

		

	except:
		pass

def division(indx,clr):
	global indx_name
	if indx==1:
		indx_name='TMAXmean'
	if indx==2:
		indx_name='TMINmean'
	if indx==3:
		indx_name='TN10p'
	if indx==4:
		indx_name='TN90p'
	if indx==5:
		indx_name='TX10p'
	if indx==6:
		indx_name='TX90p'
	if indx==7:
		indx_name='TNn'
	if indx==8:
		indx_name='TNx'
	if indx==9:
		indx_name='TXn'
	if indx==10:
		indx_name='TXx'
	division_2(indx_name,clr)


def fn1(event):
	global indx,clr
	#my_canvas.create_image(500/2,681/2,image=bd_2)
	if event.keysym=='Left':
		indx=indx-1
		if indx<1 or indx>10:
			indx=1
	if event.keysym=='Right':
		indx=indx+1
		if indx<1 or indx>10:
			indx=1
	if event.keysym=='Up':
		clr=clr+1
		if clr<1 or clr>5:
			clr=1
	if event.keysym=='Down':
		clr=clr-1
		if clr<1 or clr>5:
			clr=1
	print(f'index ={indx}   clr ={clr}')
	division(indx,clr)

indx=2
clr=2
root.bind('<Key>',fn1)
root.mainloop()